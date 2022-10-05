import os.path
from zipfile import ZipFile
from PyPDF2 import PdfReader
import csv
from openpyxl import load_workbook


def test_docs_zip():
    pdf_file = os.path.abspath('./test_docs1/resources/izveschenie.pdf')
    csv_file = os.path.abspath('./test_docs1/resources/industry.csv')
    xlsx_file = os.path.abspath('./test_docs1/resources/arenda.xlsx')

    print(pdf_file)
    print(csv_file)
    print(xlsx_file)

    archive = ZipFile('Archive.zip', 'w')
    archive.write(xlsx_file, arcname='xlsx_put.xlsx')
    archive.write(csv_file, arcname='csv_put.csv')
    archive.write(pdf_file, arcname='pdf_put.pdf')

    archive.close()


def test_read_pdf():
    pdf_read = PdfReader('./test_docs1/resources/izveschenie.pdf')
    amount_pages = len(pdf_read.pages)
    assert amount_pages == 5
    page = pdf_read.pages[1]
    text = page.extract_text()
    assert 'ПЕРИОД' in text


def test_read_csv():
    with open('./test_docs1/resources/industry.csv') as csvfile:
        tablitsa = csv.reader(csvfile)
        for line_no, line in enumerate(tablitsa, 1):
            if line_no == 2:
                assert 'Level' in line[1]


def test_read_xlsx():
    workbook = load_workbook('./test_docs1/resources/arenda.xlsx')
    sheet = workbook.active
    assert 3 == sheet.cell(row=10, column=1).value
